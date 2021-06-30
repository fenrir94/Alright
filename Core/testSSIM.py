from skimage.measure import compare_ssim
import argparse
import imutils
import cv2
from openpyxl import Workbook
from openpyxl import load_workbook
import os

# ap = argparse.ArgumentParser()
# ap.add_argument("-f", "--first", required=True, help="first input image")
# ap.add_argument("-f", "--second", required=True, help="second")
# args=vars(ap.parse_args())

# imageA = cv2.imread(args["first"])
# imageB = cv2.imread(args["second"])
#
# grayA = cv2.cvtColor(imageA, cv2.COLOR_BGR2GRAY)
# grayB = cv2.cvtColor(imageB, cv2.COLOR_BGR2GRAY)
#
# (score, diff) = compare_ssim(grayA, grayB, full=True)
# diff = (diff*255).astype("uint8")
# print("SSIM: {}".format(score))


def calculateSSIM( imageAugmented, imageAugmentedName, imageOrigin):
    grayAugmented = cv2.cvtColor(imageAugmented, cv2.COLOR_BGR2GRAY)
    grayOrigin = cv2.cvtColor(imageOrigin, cv2.COLOR_BGR2GRAY)
    score, _ = compare_ssim(grayAugmented, grayOrigin, full=True)
    print(imageAugmentedName+ " SSIM: {}".format(score))

    return score


def saveSSIM( sheetname, imagesAugmented, imageAugmentedName, imageOrigin):
    # wb = Workbook()
    if os.isfile("SSIM.xlsx"):
        wb = load_workbook("SSIM.xlsx", data_only=True)
    else:
        wb = Workbook()
    ws = wb.create_sheet()
    ws.title = sheetname

    for index, imageAugmented in enumerate(imagesAugmented):
        grayAugmented = cv2.cvtColor(imageAugmented, cv2.COLOR_BGR2GRAY)
        grayOrigin = cv2.cvtColor(imageOrigin, cv2.COLOR_BGR2GRAY)
        score, _ = compare_ssim(grayAugmented, grayOrigin, full=True)
        print(str(imageAugmentedName) + " SSIM: " + str(score))
        ws.cell(index+1, 1, imageAugmentedName[index])
        ws.cell(index+1, 2, score)

    wb.save(filename="SSIM.xlsx")

